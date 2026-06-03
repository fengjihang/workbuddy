import os
import uuid
from pathlib import Path
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Form
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from ..database import get_db
from ..models.tender import Tender
from ..models.bid import Bid
from ..models.compliance import ComplianceResult
from ..schemas.compliance import ComplianceItemOut
from ..services.compliance import run_compliance_check
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter(prefix="/api/compliance", tags=["合规检查"])

UPLOAD_DIR = "uploads"
EXPORT_DIR = "exports"


@router.post("/check")
async def check_compliance(
    tender_file: UploadFile = File(...),
    bid_file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    # 保存招标文件
    Path(UPLOAD_DIR).mkdir(parents=True, exist_ok=True)
    tender_path = str(Path(UPLOAD_DIR) / f"compliance_tender_{uuid.uuid4().hex}.docx")
    tender_content = await tender_file.read()
    with open(tender_path, "wb") as f:
        f.write(tender_content)

    # 保存标书
    bid_path = str(Path(UPLOAD_DIR) / f"compliance_bid_{uuid.uuid4().hex}.docx")
    bid_content = await bid_file.read()
    with open(bid_path, "wb") as f:
        f.write(bid_content)

    # 创建临时记录
    tender = Tender(filename=tender_file.filename, file_path=tender_path,
                    file_type="docx", file_size=len(tender_content), status="已上传")
    bid = Bid(name=f"合规检查-{tender_file.filename}")
    db.add(tender)
    db.add(bid)
    db.commit()

    async def stream():
        results_buf = []
        async for msg in run_compliance_check(tender.id, tender_path, bid_path):
            import json
            try:
                data = json.loads(msg.strip())
                if data.get("type") == "item_result":
                    results_buf.append(data["data"])
                elif data.get("type") == "summary":
                    # 入库
                    for item in results_buf:
                        db.add(ComplianceResult(
                            tender_id=tender.id, bid_id=bid.id,
                            item_index=item["item_index"], item_desc=item["item_desc"],
                            category=item["category"], risk_level=item["risk_level"],
                            page_ref=item.get("page_ref", ""), status=item["status"],
                            remark=item.get("remark", ""),
                        ))
                    db.commit()
            except json.JSONDecodeError:
                pass
            yield f"data: {msg.strip()}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


@router.get("/{bid_id}/results", response_model=list[ComplianceItemOut])
def get_results(bid_id: int, db: Session = Depends(get_db)):
    return db.query(ComplianceResult).filter(
        ComplianceResult.bid_id == bid_id
    ).order_by(ComplianceResult.item_index).all()


@router.get("/{bid_id}/export")
def export_excel(bid_id: int, db: Session = Depends(get_db)):
    results = db.query(ComplianceResult).filter(
        ComplianceResult.bid_id == bid_id
    ).order_by(ComplianceResult.item_index).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "合规检查结果"

    headers = ["序号", "检查项描述", "类别", "优先级", "页码", "状态", "备注"]
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 风险等级颜色
    risk_colors = {"严重": "FF0000", "高": "FF8C00", "中": "FFD700", "低": "90EE90"}

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r.item_index).border = thin_border
        ws.cell(row=row_idx, column=2, value=r.item_desc).border = thin_border
        ws.cell(row=row_idx, column=3, value=r.category).border = thin_border
        risk_cell = ws.cell(row=row_idx, column=4, value=r.risk_level)
        risk_cell.border = thin_border
        if r.risk_level in risk_colors:
            risk_cell.font = Font(color=risk_colors[r.risk_level], bold=True)
        ws.cell(row=row_idx, column=5, value=r.page_ref or "").border = thin_border
        ws.cell(row=row_idx, column=6, value=r.status).border = thin_border
        ws.cell(row=row_idx, column=7, value=r.remark or "").border = thin_border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 30

    Path(EXPORT_DIR).mkdir(parents=True, exist_ok=True)
    filepath = str(Path(EXPORT_DIR) / f"compliance_{bid_id}.xlsx")
    wb.save(filepath)

    return FileResponse(filepath, filename="合规检查结果.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
