"""模块 10 标书检查清单 → Excel 导出"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generate_checklist_excel(module_10_content: str, export_dir: str) -> str:
    """将模块 10 的 Markdown/文本内容解析为 Excel 清单"""
    wb = Workbook()
    ws = wb.active
    ws.title = "标书检查清单"

    # 表头
    headers = ["序号", "检查项描述", "类别", "状态", "页码", "备注", "优先级", "检查时间"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 解析模块内容中的表格行
    items = _parse_checklist_items(module_10_content)

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        ws.cell(row=row_idx, column=2, value=item.get("desc", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=item.get("category", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value="").border = thin_border  # 状态留空
        ws.cell(row=row_idx, column=5, value="").border = thin_border  # 页码留空
        ws.cell(row=row_idx, column=6, value=item.get("remark", "")).border = thin_border
        ws.cell(row=row_idx, column=7, value=item.get("priority", "")).border = thin_border
        ws.cell(row=row_idx, column=8, value="").border = thin_border  # 检查时间留空

    # 列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 14

    Path(export_dir).mkdir(parents=True, exist_ok=True)
    filepath = str(Path(export_dir) / "checklist.xlsx")
    wb.save(filepath)
    return filepath


def _parse_checklist_items(content: str) -> list[dict]:
    """解析 LLM 返回的检查清单内容为结构化列表"""
    items = []
    lines = content.strip().split("\n")
    for line in lines:
        line = line.strip()
        if not line or line.startswith("#") or line.startswith("序号"):
            continue
        # 尝试按 | 分割（Markdown 表格行）
        if "|" in line:
            parts = [p.strip() for p in line.split("|")]
            parts = [p for p in parts if p]  # 去空
            if len(parts) >= 3:
                items.append({
                    "desc": parts[1] if len(parts) > 1 else "",
                    "category": parts[2] if len(parts) > 2 else "",
                    "priority": parts[3] if len(parts) > 3 else "",
                    "remark": parts[4] if len(parts) > 4 else "",
                })
        elif line[0].isdigit():
            # 尝试解析 "1. xxx" 格式
            desc = line.split(".", 1)[-1].strip() if "." in line else line
            items.append({"desc": desc, "category": "", "priority": "", "remark": ""})

    return items
